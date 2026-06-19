import io
from datetime import timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.db.models import Count, Avg
from django.db import transaction
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    Exam, Subject, CurrentAffairsCategory, MCQ, PastPaper, Syllabus,
    JobListing, Student, TestResult, ActivityLog, ContactMessage, Announcement,
    SectionContent,
    ServicePlan, ApplicationRequest, UserProfile, UserEducation, UserExperience, UserDocument,
    Post, Comment, Category, Tag, NewsSubscriber,
    AIUsage, ChatSession, ChatMessage, AISubscription,
)
from .forms import (
    JobListingForm, SyllabusForm, PastPaperForm, MCQForm,
    CurrentAffairsCategoryForm, AnnouncementForm, ExamForm, SectionContentForm,
    SubjectForm, PostForm,
)
from .mcq_parser import parse_mcq_text
from .utils import scraper_control


def dashboard_login(request):
    """Custom login page for the admin dashboard."""
    if request.user.is_authenticated and request.user.is_staff:
        return redirect('dashboard_home')

    if request.method == 'POST':
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_staff:
            login(request, user)
            return redirect('dashboard_home')
        else:
            messages.error(request, 'Invalid credentials or insufficient permissions.')

    return render(request, 'dashboard/login.html')


def dashboard_logout(request):
    logout(request)
    return redirect('dashboard_login')


@login_required(login_url='/dashboard/login/')
def dashboard_home(request):
    """Main dashboard page with stats, recent MCQs, activity, quick actions."""
    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday_start = today_start - timedelta(days=1)
    week_start = today_start - timedelta(days=7)

    # Stats
    total_mcqs = MCQ.objects.count()
    mcqs_this_week = MCQ.objects.filter(created_at__gte=week_start).count()
    registered_users = User.objects.count()
    users_this_week = User.objects.filter(date_joined__gte=week_start).count()
    active_jobs = JobListing.objects.filter(status='active').count()
    jobs_today = JobListing.objects.filter(created_at__gte=today_start).count()
    tests_today = TestResult.objects.filter(created_at__gte=today_start).count()
    tests_yesterday = TestResult.objects.filter(
        created_at__gte=yesterday_start, created_at__lt=today_start
    ).count()
    tests_vs_yesterday = tests_today - tests_yesterday

    # Recent MCQs
    recent_mcqs = MCQ.objects.select_related('exam', 'subject').all()[:5]

    # Recent Activity
    recent_activity = ActivityLog.objects.all()[:5]

    # Service stats
    service_requests_count = ApplicationRequest.objects.count()
    job_profiles_count = UserProfile.objects.count()
    service_plans_count = ServicePlan.objects.count()

    # Blog stats
    blog_posts_count = Post.objects.count()
    blog_comments_count = Comment.objects.count()
    news_subscribers_count = NewsSubscriber.objects.count()

    context = {
        'total_mcqs': f'{total_mcqs:,}',
        'mcqs_this_week': mcqs_this_week,
        'registered_users': f'{registered_users:,}',
        'users_this_week': users_this_week,
        'active_jobs': f'{active_jobs:,}',
        'jobs_today': jobs_today,
        'tests_today': f'{tests_today:,}',
        'tests_vs_yesterday': tests_vs_yesterday,
        'recent_mcqs': recent_mcqs,
        'recent_activity': recent_activity,
        'service_requests_count': service_requests_count,
        'job_profiles_count': job_profiles_count,
        'service_plans_count': service_plans_count,
        'blog_posts_count': blog_posts_count,
        'blog_comments_count': blog_comments_count,
        'news_subscribers_count': news_subscribers_count,
        'active_page': 'dashboard',
    }
    return render(request, 'dashboard/home.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_mcqs(request):
    """MCQ bank listing page."""
    mcqs = MCQ.objects.select_related('exam', 'subject', 'current_affairs_category').all()
    exams = Exam.objects.all()
    subjects = Subject.objects.annotate(mcq_count=Count('mcqs')).order_by('name')
    current_affairs_categories = (
        CurrentAffairsCategory.objects
        .annotate(mcq_count=Count('mcqs'))
        .order_by('region', 'sort_order', 'name')
    )

    # Filters
    exam_filter = request.GET.get('exam')
    subject_filter = request.GET.get('subject')
    ca_category_filter = request.GET.get('current_affairs_category')
    status_filter = request.GET.get('status')

    if exam_filter:
        mcqs = mcqs.filter(exam__slug=exam_filter)
    if subject_filter:
        mcqs = mcqs.filter(subject__slug=subject_filter)
    if ca_category_filter:
        mcqs = mcqs.filter(current_affairs_category__slug=ca_category_filter)
    if status_filter:
        mcqs = mcqs.filter(status=status_filter)

    total_count = mcqs.count()

    # Pagination
    paginator = Paginator(mcqs, 25) # 25 items per page
    page = request.GET.get('page')
    try:
        mcqs_page = paginator.page(page)
    except PageNotAnInteger:
        mcqs_page = paginator.page(1)
    except EmptyPage:
        mcqs_page = paginator.page(paginator.num_pages)

    context = {
        'mcqs': mcqs_page,
        'exams': exams,
        'subjects': subjects,
        'current_affairs_categories': current_affairs_categories,
        'active_page': 'mcqs',
        'total_count': total_count,
    }
    return render(request, 'dashboard/mcqs.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_mcq_create(request):
    """MCQ creation and Excel upload page."""
    exams = Exam.objects.all()
    subjects = Subject.objects.all()
    current_affairs_categories = list(
        CurrentAffairsCategory.objects
        .filter(is_active=True)
        .order_by('region', 'sort_order', 'name')
        .values('name', 'slug', 'region')
    )
    context = {
        'active_page': 'mcqs',
        'exams': exams,
        'subjects': subjects,
        'current_affairs_categories': current_affairs_categories,
    }
    return render(request, 'dashboard/mcq_create.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_mcq_create_with_paper(request, paper_pk):
    """Create an MCQ pre-linked to a specific past paper."""
    paper = get_object_or_404(PastPaper, pk=paper_pk)
    if request.method == 'POST':
        form = MCQForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'MCQ linked to paper successfully.')
            return redirect('dashboard_past_papers')
    else:
        form = MCQForm(initial={
            'past_paper': paper.pk,
            'exam': paper.exam.pk if paper.exam else None,
            'subject': paper.subject.pk if paper.subject else None,
        })

    context = {
        'form': form,
        'title': f'Add MCQ to {paper.title}',
        'active_page': 'past_papers',
    }
    return render(request, 'dashboard/form.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_paper_mcqs(request, paper_pk):
    """Manage all MCQs linked to a specific past paper — add, view, edit, delete."""
    paper = get_object_or_404(PastPaper.objects.select_related('exam', 'subject'), pk=paper_pk)

    if request.method == 'POST':
        form = MCQForm(request.POST)
        if form.is_valid():
            mcq = form.save(commit=False)
            mcq.past_paper = paper
            mcq.created_by = request.user
            mcq.save()
            messages.success(request, 'MCQ added to paper successfully.')
            return redirect('dashboard_paper_mcqs', paper_pk=paper.pk)
    else:
        form = MCQForm(initial={
            'past_paper': paper.pk,
            'exam': paper.exam.pk if paper.exam else None,
            'subject': paper.subject.pk if paper.subject else None,
        })

    mcqs = MCQ.objects.filter(past_paper=paper).select_related('exam', 'subject').order_by('id')
    total_count = mcqs.count()

    # Pagination
    paginator = Paginator(mcqs, 25)
    page = request.GET.get('page')
    try:
        mcqs_page = paginator.page(page)
    except PageNotAnInteger:
        mcqs_page = paginator.page(1)
    except EmptyPage:
        mcqs_page = paginator.page(paginator.num_pages)

    context = {
        'paper': paper,
        'form': form,
        'mcqs': mcqs_page,
        'total_count': total_count,
        'active_page': 'past_papers',
    }
    return render(request, 'dashboard/paper_mcqs.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_mcq_bulk_upload(request):
    """Import MCQs into the dashboard from pasted text."""
    exams = Exam.objects.all()
    subjects = Subject.objects.all()
    current_affairs_categories = list(
        CurrentAffairsCategory.objects
        .filter(is_active=True)
        .order_by('region', 'sort_order', 'name')
        .values('name', 'slug', 'region')
    )

    context = {
        'active_page': 'mcqs',
        'exams': exams,
        'subjects': subjects,
        'current_affairs_categories': current_affairs_categories,
        'status': 'draft',
    }

    if request.method == 'POST':
        action = request.POST.get('action', 'preview')
        exam_id = request.POST.get('exam')
        subject_id = request.POST.get('subject')
        raw_text = request.POST.get('raw_text', '').strip()
        status = request.POST.get('status', 'draft')
        source_url = request.POST.get('source_url', '').strip()
        category_slug = request.POST.get('current_affairs_category', '').strip()

        context.update({
            'raw_text': raw_text,
            'source_url': source_url,
            'status': status,
            'selected_exam': exam_id,
            'selected_subject': subject_id,
            'selected_current_affairs_category': category_slug,
        })

        if status not in ('draft', 'published', 'flagged'):
            status = 'draft'

        exam = Exam.objects.filter(pk=exam_id).first()
        subject = Subject.objects.filter(pk=subject_id).first()
        if not exam or not subject:
            messages.error(request, 'Please select a valid exam and subject.')
            return render(request, 'dashboard/mcq_bulk_upload.html', context)
        current_affairs_category = None
        if subject.slug == 'current-affairs' and category_slug:
            current_affairs_category = CurrentAffairsCategory.objects.filter(
                slug=category_slug,
                is_active=True,
            ).first()

        text = raw_text

        if not text:
            messages.error(request, 'Please paste MCQ text.')
            return render(request, 'dashboard/mcq_bulk_upload.html', context)

        parsed = parse_mcq_text(text)
        context['raw_text'] = text
        context['parsed_mcqs'] = parsed
        context['parsed_count'] = len(parsed)

        if not parsed:
            messages.warning(request, 'No MCQs could be parsed. Please check the format and try again.')
            return render(request, 'dashboard/mcq_bulk_upload.html', context)

        if action != 'save':
            messages.info(request, f'Extracted text and parsed {len(parsed)} MCQs. Review or edit before saving.')
            return render(request, 'dashboard/mcq_bulk_upload.html', context)

        created_count = 0
        with transaction.atomic():
            for item in parsed:
                MCQ.objects.create(
                    question_text=item['question_text'],
                    option_a=item['option_a'],
                    option_b=item['option_b'],
                    option_c=item['option_c'],
                    option_d=item['option_d'],
                    correct_option=item['correct_option'] or 'A',
                    explanation=item['explanation'],
                    exam=exam,
                    subject=subject,
                    current_affairs_category=current_affairs_category,
                    status=status,
                    source_url=source_url,
                    created_by=request.user,
                )
                created_count += 1

        messages.success(request, f'Successfully imported {created_count} MCQs.')
        return redirect('dashboard_mcqs')

    return render(request, 'dashboard/mcq_bulk_upload.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_mcq_edit(request, pk):
    mcq = get_object_or_404(MCQ, pk=pk)
    if request.method == 'POST':
        form = MCQForm(request.POST, instance=mcq)
        if form.is_valid():
            form.save()
            messages.success(request, 'MCQ updated successfully!')
            return redirect('dashboard_mcqs')
    else:
        form = MCQForm(instance=mcq)
    
    context = {
        'form': form,
        'title': 'Edit MCQ',
        'active_page': 'mcqs'
    }
    return render(request, 'dashboard/form.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_current_affairs_categories(request):
    """Create and manage Current Affairs category buckets."""
    categories = (
        CurrentAffairsCategory.objects
        .annotate(mcq_count=Count('mcqs'))
        .order_by('region', 'sort_order', 'name')
    )

    region_filter = request.GET.get('region')
    status_filter = request.GET.get('status')

    if region_filter in ('pakistan', 'world'):
        categories = categories.filter(region=region_filter)
    if status_filter == 'active':
        categories = categories.filter(is_active=True)
    elif status_filter == 'inactive':
        categories = categories.filter(is_active=False)

    if request.method == 'POST':
        form = CurrentAffairsCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Current Affairs category added successfully.')
            return redirect('dashboard_current_affairs_categories')
    else:
        form = CurrentAffairsCategoryForm()

    context = {
        'active_page': 'ca_categories',
        'categories': categories,
        'form': form,
        'total_count': categories.count(),
    }
    return render(request, 'dashboard/current_affairs_categories.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_current_affairs_category_edit(request, pk):
    category = get_object_or_404(CurrentAffairsCategory, pk=pk)
    if request.method == 'POST':
        form = CurrentAffairsCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Current Affairs category updated successfully.')
            return redirect('dashboard_current_affairs_categories')
    else:
        form = CurrentAffairsCategoryForm(instance=category)

    context = {
        'form': form,
        'title': 'Edit Current Affairs Category',
        'active_page': 'ca_categories',
    }
    return render(request, 'dashboard/form.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_current_affairs_category_delete(request, pk):
    category = get_object_or_404(CurrentAffairsCategory, pk=pk)
    if request.method == 'POST':
        name = category.name
        mcq_count = category.mcqs.count()
        category.delete()
        if mcq_count:
            messages.success(
                request,
                f'Deleted "{name}" and removed it from {mcq_count} linked MCQs.'
            )
        else:
            messages.success(request, f'Deleted "{name}".')
    return redirect('dashboard_current_affairs_categories')


@login_required(login_url='/dashboard/login/')
def dashboard_past_papers(request):
    """Past papers listing page."""
    papers = PastPaper.objects.select_related('exam', 'subject').all()
    total_count = papers.count()

    # Pagination
    paginator = Paginator(papers, 25)
    page = request.GET.get('page')
    try:
        papers_page = paginator.page(page)
    except PageNotAnInteger:
        papers_page = paginator.page(1)
    except EmptyPage:
        papers_page = paginator.page(paginator.num_pages)

    context = {
        'papers': papers_page,
        'active_page': 'past_papers',
        'total_count': total_count,
    }
    return render(request, 'dashboard/past_papers.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_syllabus(request):
    """Syllabus listing page."""
    syllabi = Syllabus.objects.select_related('exam').all()
    total_count = syllabi.count()

    # Pagination
    paginator = Paginator(syllabi, 25)
    page = request.GET.get('page')
    try:
        syllabi_page = paginator.page(page)
    except PageNotAnInteger:
        syllabi_page = paginator.page(1)
    except EmptyPage:
        syllabi_page = paginator.page(paginator.num_pages)

    context = {
        'syllabi': syllabi_page,
        'active_page': 'syllabus',
        'total_count': total_count,
    }
    return render(request, 'dashboard/syllabus.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_jobs(request):
    """Job listings page."""
    jobs = JobListing.objects.select_related('exam', 'syllabus').all()
    total_count = jobs.count()

    # Pagination
    paginator = Paginator(jobs, 25)
    page = request.GET.get('page')
    try:
        jobs_page = paginator.page(page)
    except PageNotAnInteger:
        jobs_page = paginator.page(1)
    except EmptyPage:
        jobs_page = paginator.page(paginator.num_pages)

    context = {
        'jobs': jobs_page,
        'active_page': 'jobs',
        'total_count': total_count,
    }
    return render(request, 'dashboard/jobs.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_students(request):
    """Students listing page."""
    students = Student.objects.select_related('user').all()
    total_count = students.count()

    # Pagination
    paginator = Paginator(students, 25)
    page = request.GET.get('page')
    try:
        students_page = paginator.page(page)
    except PageNotAnInteger:
        students_page = paginator.page(1)
    except EmptyPage:
        students_page = paginator.page(paginator.num_pages)

    context = {
        'students': students_page,
        'active_page': 'students',
        'total_count': total_count,
    }
    return render(request, 'dashboard/students.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_test_results(request):
    """Test results listing page."""
    results = TestResult.objects.select_related('student', 'exam', 'subject').all()
    total_count = results.count()

    # Pagination
    paginator = Paginator(results, 25)
    page = request.GET.get('page')
    try:
        results_page = paginator.page(page)
    except PageNotAnInteger:
        results_page = paginator.page(1)
    except EmptyPage:
        results_page = paginator.page(paginator.num_pages)

    context = {
        'results': results_page,
        'active_page': 'test_results',
        'total_count': total_count,
    }
    return render(request, 'dashboard/test_results.html', context)


# ─── Page sections CRUD ────────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_sections(request):
    """List and create editable frontend section headings."""
    if request.method == 'POST':
        form = SectionContentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Section content added successfully.')
            return redirect('dashboard_sections')
    else:
        form = SectionContentForm()

    sections = SectionContent.objects.all()
    context = {
        'active_page': 'sections',
        'sections': sections,
        'form': form,
        'total_count': sections.count(),
    }
    return render(request, 'dashboard/sections.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_section_edit(request, pk):
    section = get_object_or_404(SectionContent, pk=pk)
    if request.method == 'POST':
        form = SectionContentForm(request.POST, instance=section)
        if form.is_valid():
            form.save()
            messages.success(request, 'Section content updated successfully.')
            return redirect('dashboard_sections')
    else:
        form = SectionContentForm(instance=section)

    context = {
        'form': form,
        'title': 'Edit Section Content',
        'active_page': 'sections',
    }
    return render(request, 'dashboard/form.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_section_delete(request, pk):
    section = get_object_or_404(SectionContent, pk=pk)
    if request.method == 'POST':
        section.delete()
        messages.success(request, 'Section content deleted.')
    return redirect('dashboard_sections')


# ─── Exam boards CRUD ─────────────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_exams(request):
    """List and create exam boards."""
    if request.method == 'POST':
        form = ExamForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Exam board added successfully.')
            return redirect('dashboard_exams')
    else:
        form = ExamForm()

    exams = Exam.objects.annotate(
        mcq_count=Count('mcqs', distinct=True),
        paper_count=Count('past_papers', distinct=True),
        syllabus_count=Count('syllabi', distinct=True),
    ).order_by('name')
    context = {
        'active_page': 'exams',
        'exams': exams,
        'form': form,
        'total_count': exams.count(),
    }
    return render(request, 'dashboard/exams.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_exam_edit(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    if request.method == 'POST':
        form = ExamForm(request.POST, instance=exam)
        if form.is_valid():
            form.save()
            messages.success(request, 'Exam board updated successfully.')
            return redirect('dashboard_exams')
    else:
        form = ExamForm(instance=exam)

    context = {
        'form': form,
        'title': 'Edit Exam Board',
        'active_page': 'exams',
    }
    return render(request, 'dashboard/form.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_exam_delete(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    if request.method == 'POST':
        name = exam.name
        exam.delete()
        messages.success(request, f'Deleted "{name}" and its linked content.')
    return redirect('dashboard_exams')


# ─── Subjects CRUD ────────────────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_subjects(request):
    """List and create subjects."""
    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Subject added successfully.')
            return redirect('dashboard_subjects')
    else:
        form = SubjectForm()

    subjects = Subject.objects.annotate(mcq_count=Count('mcqs', distinct=True)).order_by('name')
    context = {
        'active_page': 'subjects',
        'subjects': subjects,
        'form': form,
        'total_count': subjects.count(),
    }
    return render(request, 'dashboard/subjects.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_subject_edit(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            messages.success(request, 'Subject updated successfully.')
            return redirect('dashboard_subjects')
    else:
        form = SubjectForm(instance=subject)

    context = {
        'form': form,
        'title': 'Edit Subject',
        'active_page': 'subjects',
    }
    return render(request, 'dashboard/form.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_subject_delete(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        name = subject.name
        mcq_count = subject.mcqs.count()
        subject.delete()
        messages.success(request, f'Deleted "{name}" and its {mcq_count} linked MCQs.')
    return redirect('dashboard_subjects')


# ─── Announcements CRUD ───────────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_announcements(request):
    """List and create announcement-bar items."""
    if request.method == 'POST':
        form = AnnouncementForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Announcement added successfully.')
            return redirect('dashboard_announcements')
    else:
        form = AnnouncementForm()

    announcements = Announcement.objects.all()
    context = {
        'active_page': 'announcements',
        'announcements': announcements,
        'form': form,
        'total_count': announcements.count(),
    }
    return render(request, 'dashboard/announcements.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_announcement_edit(request, pk):
    announcement = get_object_or_404(Announcement, pk=pk)
    if request.method == 'POST':
        form = AnnouncementForm(request.POST, instance=announcement)
        if form.is_valid():
            form.save()
            messages.success(request, 'Announcement updated successfully.')
            return redirect('dashboard_announcements')
    else:
        form = AnnouncementForm(instance=announcement)

    context = {
        'form': form,
        'title': 'Edit Announcement',
        'active_page': 'announcements',
    }
    return render(request, 'dashboard/form.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_announcement_delete(request, pk):
    announcement = get_object_or_404(Announcement, pk=pk)
    if request.method == 'POST':
        announcement.delete()
        messages.success(request, 'Announcement deleted.')
    return redirect('dashboard_announcements')


@login_required(login_url='/dashboard/login/')
def dashboard_settings(request):
    """Settings page placeholder."""
    context = {'active_page': 'settings'}
    return render(request, 'dashboard/settings.html', context)


# ─── Job listings CRUD ────────────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_job_create(request):
    if request.method == 'POST':
        form = JobListingForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Job listing posted successfully!')
            return redirect('dashboard_jobs')
    else:
        form = JobListingForm()
    
    context = {
        'form': form,
        'title': 'Post Job',
        'active_page': 'jobs'
    }
    return render(request, 'dashboard/form.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_job_edit(request, pk):
    job = get_object_or_404(JobListing, pk=pk)
    if request.method == 'POST':
        form = JobListingForm(request.POST, instance=job)
        if form.is_valid():
            form.save()
            messages.success(request, 'Job listing updated successfully!')
            return redirect('dashboard_jobs')
    else:
        form = JobListingForm(instance=job)
    
    context = {
        'form': form,
        'title': 'Edit Job Listing',
        'active_page': 'jobs'
    }
    return render(request, 'dashboard/form.html', context)


# ─── Syllabus CRUD ────────────────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_syllabus_create(request):
    if request.method == 'POST':
        form = SyllabusForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Syllabus entry created successfully!')
            return redirect('dashboard_syllabus')
    else:
        form = SyllabusForm()
    
    context = {
        'form': form,
        'title': 'Add Syllabus Entry',
        'active_page': 'syllabus'
    }
    return render(request, 'dashboard/form.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_syllabus_edit(request, pk):
    syllabus = get_object_or_404(Syllabus, pk=pk)
    if request.method == 'POST':
        form = SyllabusForm(request.POST, request.FILES, instance=syllabus)
        if form.is_valid():
            form.save()
            messages.success(request, 'Syllabus entry updated successfully!')
            return redirect('dashboard_syllabus')
    else:
        form = SyllabusForm(instance=syllabus)
    
    context = {
        'form': form,
        'title': 'Edit Syllabus Entry',
        'active_page': 'syllabus'
    }
    return render(request, 'dashboard/form.html', context)


# ─── Past Papers CRUD ─────────────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_past_paper_create(request):
    if request.method == 'POST':
        form = PastPaperForm(request.POST, request.FILES)
        if form.is_valid():
            paper = form.save(commit=False)
            paper.created_by = request.user
            paper.save()
            messages.success(request, 'Past paper uploaded successfully!')
            return redirect('dashboard_past_papers')
    else:
        form = PastPaperForm()
    
    context = {
        'form': form,
        'title': 'Upload Past Paper',
        'active_page': 'past_papers'
    }
    return render(request, 'dashboard/form.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_past_paper_edit(request, pk):
    paper = get_object_or_404(PastPaper, pk=pk)
    if request.method == 'POST':
        form = PastPaperForm(request.POST, request.FILES, instance=paper)
        if form.is_valid():
            form.save()
            messages.success(request, 'Past paper updated successfully!')
            return redirect('dashboard_past_papers')
    else:
        form = PastPaperForm(instance=paper)
    
    context = {
        'form': form,
        'title': 'Edit Past Paper',
        'active_page': 'past_papers'
    }
    return render(request, 'dashboard/form.html', context)


# Global state for simple log streaming without redis/celery
SCRAPER_STATE = {
    'running': False,
    'stop_requested': False,
    'logs': []
}

def append_scraper_log(line):
    if str(line).strip():
        SCRAPER_STATE['logs'].append(str(line).strip())
    if len(SCRAPER_STATE['logs']) > 1000:
        SCRAPER_STATE['logs'] = SCRAPER_STATE['logs'][-1000:]

class ScraperLogStream:
    """A stream wrapper to capture call_command output line by line."""
    def write(self, s):
        if s.strip('\r\n'):
            for line in s.strip('\r\n').split('\n'):
                if line.strip():
                    append_scraper_log(line)
    def flush(self):
        pass

@login_required(login_url='/dashboard/login/')
def dashboard_scraper(request):
    """Scraper control panel — supports scrape_data, scrape_testpoint, scrape_pastpapers."""
    from django.core.management import call_command
    import threading

    if request.method == 'POST':
        action = request.POST.get('action', 'start')

        if action == 'stop':
            if SCRAPER_STATE['running']:
                scraper_control.request_stop()
                SCRAPER_STATE['stop_requested'] = True
                append_scraper_log('[STOP] Stop requested from dashboard. Saving collected data before exit...')
                messages.warning(request, 'Stop requested. The scraper will save collected data before exiting.')
            else:
                messages.info(request, 'No scraper is currently running.')
            return redirect('dashboard_scraper')

        scraper_cmd = request.POST.get('scraper_cmd', 'scrape_data')

        if not SCRAPER_STATE['running']:
            scraper_control.clear_stop()
            SCRAPER_STATE['running'] = True
            SCRAPER_STATE['stop_requested'] = False
            SCRAPER_STATE['logs'] = []

            # Capture ALL POST values before the thread starts — the request
            # object is not safe to read from inside a background thread after
            # the HTTP response has been sent.
            _cmd        = scraper_cmd
            _sd_type    = request.POST.get('scrape_type', 'all')
            _sd_start   = int(request.POST.get('start_page', 1))
            _sd_max     = int(request.POST.get('max_pages', 0))
            _tp_mode    = request.POST.get('tp_mode', 'past-papers')
            _tp_exam    = request.POST.get('tp_exam', 'all')
            _tp_engine  = request.POST.get('tp_engine', 'curl')
            _tp_max     = int(request.POST.get('tp_max_papers', 0))
            _tp_max_pages = int(request.POST.get('tp_max_pages', 0))
            _tp_subj    = request.POST.get('tp_subject', '')
            _tp_debug   = request.POST.get('tp_debug') == '1'
            _gt_subject = request.POST.get('gt_subject', '')
            _gt_max     = int(request.POST.get('gt_max_tests', 0))
            _gt_debug   = request.POST.get('gt_debug') == '1'
            _gt_dry_run = request.POST.get('gt_dry_run') == '1'
            _pp_exam    = request.POST.get('pp_exam', 'all')
            _pp_engine  = request.POST.get('pp_engine', 'curl')
            _pp_max     = int(request.POST.get('pp_max_posts', 0))
            _pp_subj    = request.POST.get('pp_subject', '')
            _pp_debug   = request.POST.get('pp_debug') == '1'

            def run_scrape():
                stream = ScraperLogStream()
                try:
                    if _cmd == 'scrape_data':
                        kwargs = {'type': _sd_type, 'start_page': _sd_start}
                        if _sd_max > 0:
                            kwargs['max_pages'] = _sd_max
                        call_command('scrape_data', stdout=stream, stderr=stream, **kwargs)

                    elif _cmd == 'scrape_testpoint':
                        kwargs = {
                            'mode': _tp_mode,
                            'engine': _tp_engine,
                            'subject': _tp_subj,
                            'debug': _tp_debug,
                        }
                        if _tp_mode == 'past-papers':
                            kwargs['exam'] = _tp_exam
                            if _tp_max > 0:
                                kwargs['max_papers'] = _tp_max
                        else:
                            if _tp_max_pages > 0:
                                kwargs['max_pages'] = _tp_max_pages
                        call_command(
                            'scrape_testpoint_v2', stdout=stream, stderr=stream, **kwargs
                        )

                    elif _cmd == 'scrape_gotest':
                        call_command(
                            'scrape_gotest', stdout=stream, stderr=stream,
                            subject=_gt_subject, max_tests=_gt_max,
                            debug=_gt_debug, dry_run=_gt_dry_run,
                        )

                    elif _cmd == 'scrape_pastpapers':
                        call_command(
                            'scrape_pastpapers', stdout=stream, stderr=stream,
                            exam=_pp_exam, engine=_pp_engine,
                            max_posts=_pp_max, subject=_pp_subj, debug=_pp_debug,
                        )

                except Exception as exc:
                    import traceback
                    stream.write(f'[ERROR] {exc}')
                    stream.write(traceback.format_exc())
                finally:
                    if SCRAPER_STATE.get('stop_requested'):
                        stream.write('[STOP] Scraper stopped gracefully after saving collected data.')
                    SCRAPER_STATE['running'] = False
                    SCRAPER_STATE['stop_requested'] = False
                    scraper_control.clear_stop()

            thread = threading.Thread(target=run_scrape, daemon=True)
            thread.start()
            messages.success(request, f'Scraper "{_cmd}" started in background.')
        else:
            messages.warning(request, 'A scraper is already running!')

        return redirect('dashboard_scraper')

    context = {
        'active_page':     'scraper',
        'scraper_running': SCRAPER_STATE['running'],
        'stop_requested':  SCRAPER_STATE['stop_requested'],
        'scraper_log':     SCRAPER_STATE['logs'],
        'mcq_count':       MCQ.objects.count(),
        'paper_count':     PastPaper.objects.count(),
        'job_count':       JobListing.objects.count(),
    }
    return render(request, 'dashboard/scraper.html', context)

@login_required(login_url='/dashboard/login/')
def dashboard_scraper_status(request):
    """Returns JSON of the current scraper status and logs for the live UI."""
    from django.http import JsonResponse
    return JsonResponse({
        'running': SCRAPER_STATE['running'],
        'stop_requested': SCRAPER_STATE['stop_requested'],
        'logs': SCRAPER_STATE['logs']
    })

@login_required(login_url='/dashboard/login/')
def dashboard_trigger_scrape(request):
    """Legacy quick-run — kept for backward compat; redirects to scraper panel."""
    return redirect('dashboard_scraper')


@login_required(login_url='/dashboard/login/')
def dashboard_contact_messages(request):
    """List all contact form submissions."""
    status_filter = request.GET.get('status', '')
    qs = ContactMessage.objects.all()
    if status_filter:
        qs = qs.filter(status=status_filter)

    paginator = Paginator(qs, 20)
    page = request.GET.get('page')
    try:
        msgs = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        msgs = paginator.page(1)

    return render(request, 'dashboard/contact_messages.html', {
        'active_page':    'contact',
        'msgs':           msgs,
        'total_count':    qs.count(),
        'unread_count':   ContactMessage.objects.filter(status='unread').count(),
        'status_filter':  status_filter,
    })


@login_required(login_url='/dashboard/login/')
def dashboard_contact_mark_read(request, pk):
    """Mark a contact message as read."""
    msg = get_object_or_404(ContactMessage, pk=pk)
    msg.status = 'read'
    msg.save()
    return redirect('dashboard_contact_messages')


@login_required(login_url='/dashboard/login/')
def dashboard_contact_delete(request, pk):
    """Delete a contact message."""
    msg = get_object_or_404(ContactMessage, pk=pk)
    msg.delete()
    return redirect('dashboard_contact_messages')


# ─── MCQ Excel Export / Import ───────────────────────────────────────────────

HEADER = [
    'id', 'question_text', 'option_a', 'option_b', 'option_c', 'option_d',
    'correct_option', 'explanation', 'exam_slug', 'subject_slug', 'status',
]

HEADER_NOTES = [
    'DO NOT EDIT',
    'The full question text',
    'Option A text',
    'Option B text',
    'Option C text',
    'Option D text (optional)',
    'A / B / C / D',
    'Optional explanation',
    'Exam slug (e.g. ppsc) — DO NOT CHANGE',
    'Subject slug — DO NOT CHANGE',
    'draft / published / flagged',
]


@login_required(login_url='/dashboard/login/')
def dashboard_mcq_export(request):
    """Download MCQs as an Excel file. Filters: subject, exam, from_row, to_row."""
    exam_slug    = request.GET.get('exam', '')
    subject_slug = request.GET.get('subject', '')
    from_row     = request.GET.get('from_row', '')
    to_row       = request.GET.get('to_row', '')

    qs = MCQ.objects.select_related('exam', 'subject').order_by('id')
    if exam_slug:
        qs = qs.filter(exam__slug=exam_slug)
    if subject_slug:
        qs = qs.filter(subject__slug=subject_slug)

    # Convert to list for slicing by row number
    mcq_list = list(qs)
    total = len(mcq_list)

    try:
        fr = max(1, int(from_row)) - 1 if from_row else 0
        tr = int(to_row) if to_row else total
        tr = min(tr, total)
        mcq_list = mcq_list[fr:tr]
    except (ValueError, TypeError):
        pass

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MCQs'

    # Styles
    hdr_fill = PatternFill('solid', fgColor='0C3638')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    note_fill = PatternFill('solid', fgColor='E8F5E9')
    note_font = Font(italic=True, color='444444', size=9)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap   = Alignment(vertical='top', wrap_text=True)

    # Row 1 — headers
    for col_idx, h in enumerate(HEADER, 1):
        cell = ws.cell(row=1, column=col_idx, value=h.upper())
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border

    # Row 2 — notes
    for col_idx, note in enumerate(HEADER_NOTES, 1):
        cell = ws.cell(row=2, column=col_idx, value=note)
        cell.font = note_font
        cell.fill = note_fill
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30

    # Data rows
    for row_idx, mcq in enumerate(mcq_list, 3):
        row = [
            mcq.id,
            mcq.question_text,
            mcq.option_a,
            mcq.option_b,
            mcq.option_c,
            mcq.option_d,
            mcq.correct_option,
            mcq.explanation,
            mcq.exam.slug,
            mcq.subject.slug,
            mcq.status,
        ]
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = wrap
            cell.border = border

    # Column widths
    widths = [8, 60, 35, 35, 35, 35, 10, 45, 14, 20, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze top 2 rows
    ws.freeze_panes = 'A3'

    # Protect id / exam_slug / subject_slug columns with a note (col A, I, J)
    ws.protection.sheet = False  # sheet itself not locked; just visual cue via fill

    filename = f"mcqs_export_{subject_slug or exam_slug or 'all'}_{fr+1}_to_{fr+len(mcq_list)}.xlsx"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required(login_url='/dashboard/login/')
def dashboard_mcq_import(request):
    """Upload an edited Excel file and bulk-update MCQs by ID."""
    if request.method != 'POST':
        return redirect('dashboard_mcqs')

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, 'No file uploaded.')
        return redirect('dashboard_mcqs')

    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'Please upload a valid .xlsx file.')
        return redirect('dashboard_mcqs')

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
    except Exception as e:
        messages.error(request, f'Could not read Excel file: {e}')
        return redirect('dashboard_mcqs')

    # Validate header row
    headers = [str(ws.cell(1, c).value or '').lower().strip() for c in range(1, len(HEADER) + 1)]
    if headers != HEADER:
        messages.error(request, f'Invalid column headers. Expected: {", ".join(HEADER)}')
        return redirect('dashboard_mcqs')

    updated = 0
    skipped = 0
    errors  = []

    # Rows start at 3 (row 1 = headers, row 2 = notes)
    for row_num in range(3, ws.max_row + 1):
        row_vals = [ws.cell(row_num, c).value for c in range(1, len(HEADER) + 1)]
        if all(v is None for v in row_vals):
            continue  # skip fully empty rows

        try:
            mcq_id       = int(row_vals[0])
            question     = str(row_vals[1] or '').strip()
            opt_a        = str(row_vals[2] or '').strip()
            opt_b        = str(row_vals[3] or '').strip()
            opt_c        = str(row_vals[4] or '').strip()
            opt_d        = str(row_vals[5] or '').strip()
            correct      = str(row_vals[6] or '').strip().upper()
            explanation  = str(row_vals[7] or '').strip()
            status       = str(row_vals[10] or 'draft').strip().lower()
        except (ValueError, TypeError) as e:
            errors.append(f'Row {row_num}: bad data — {e}')
            skipped += 1
            continue

        if not question or correct not in ('A', 'B', 'C', 'D'):
            errors.append(f'Row {row_num}: missing question or invalid correct_option "{correct}"')
            skipped += 1
            continue

        if status not in ('draft', 'published', 'flagged'):
            status = 'draft'

        try:
            mcq = MCQ.objects.get(pk=mcq_id)
            mcq.question_text  = question
            mcq.option_a       = opt_a
            mcq.option_b       = opt_b
            mcq.option_c       = opt_c
            mcq.option_d       = opt_d
            mcq.correct_option = correct
            mcq.explanation    = explanation
            mcq.status         = status
            mcq.save()
            updated += 1
        except MCQ.DoesNotExist:
            errors.append(f'Row {row_num}: MCQ id={mcq_id} not found — skipped')
            skipped += 1

    summary = f'Import complete — {updated} updated, {skipped} skipped.'
    if errors:
        summary += f' Issues: ' + ' | '.join(errors[:5])
        if len(errors) > 5:
            summary += f' … and {len(errors)-5} more.'
        messages.warning(request, summary)
    else:
        messages.success(request, summary)

    return redirect('dashboard_mcqs')


# ─── Service Requests Dashboard ──────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_service_requests(request):
    """Admin dashboard for paid application service requests."""
    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=7)

    reqs = ApplicationRequest.objects.select_related('user', 'job', 'plan', 'assigned_to').prefetch_related('admin_notes')

    # Stats
    total_requests = reqs.count()
    pending_payment = reqs.filter(request_status='payment_pending').count()
    in_progress = reqs.filter(request_status='in_progress').count()
    submitted = reqs.filter(request_status='submitted').count()
    failed = reqs.filter(request_status='failed').count()
    refunded = reqs.filter(request_status='refunded').count()
    total_revenue = sum(
        float(r.payment_amount or 0) for r in reqs.filter(payment_status='paid')
    )
    requests_today = reqs.filter(created_at__gte=today_start).count()
    requests_this_week = reqs.filter(created_at__gte=week_start).count()

    # Filters
    status_filter = request.GET.get('status')
    payment_filter = request.GET.get('payment_status')
    if status_filter:
        reqs = reqs.filter(request_status=status_filter)
    if payment_filter:
        reqs = reqs.filter(payment_status=payment_filter)

    # Pagination
    paginator = Paginator(reqs.order_by('-created_at'), 25)
    page = request.GET.get('page')
    try:
        reqs_page = paginator.page(page)
    except PageNotAnInteger:
        reqs_page = paginator.page(1)
    except EmptyPage:
        reqs_page = paginator.page(paginator.num_pages)

    context = {
        'requests': reqs_page,
        'total_requests': total_requests,
        'pending_payment': pending_payment,
        'in_progress': in_progress,
        'submitted': submitted,
        'failed': failed,
        'refunded': refunded,
        'total_revenue': f'{total_revenue:,.0f}',
        'requests_today': requests_today,
        'requests_this_week': requests_this_week,
        'active_page': 'service_requests',
        'status_filter': status_filter,
        'payment_filter': payment_filter,
    }
    return render(request, 'dashboard/service_requests.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_service_request_detail(request, pk):
    """Admin: view single service request details."""
    req = get_object_or_404(
        ApplicationRequest.objects.select_related('user', 'job', 'plan', 'assigned_to').prefetch_related('admin_notes__added_by'),
        pk=pk
    )
    return render(request, 'dashboard/service_request_detail.html', {
        'req': req,
        'active_page': 'service_requests',
    })


# ─── Job Profiles Dashboard ──────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_job_profiles(request):
    """Admin dashboard for user job profiles."""
    profiles = UserProfile.objects.select_related('user').prefetch_related(
        'user__educations', 'user__experiences', 'user__documents'
    )

    # Stats
    total_profiles = profiles.count()
    profiles_this_week = profiles.filter(created_at__gte=timezone.now() - timedelta(days=7)).count()
    total_education = UserEducation.objects.count()
    total_experience = UserExperience.objects.count()
    total_documents = UserDocument.objects.count()

    # Profile completeness
    complete_profiles = 0
    partial_profiles = 0
    for p in profiles:
        has_personal = all([p.full_name, p.cnic, p.dob, p.phone, p.permanent_address])
        has_education = p.user.educations.exists()
        has_experience = p.user.experiences.exists()
        if has_personal and has_education and has_experience:
            complete_profiles += 1
        elif has_personal or has_education or has_experience:
            partial_profiles += 1

    incomplete_profiles = total_profiles - complete_profiles - partial_profiles

    # Pagination
    paginator = Paginator(profiles.order_by('-updated_at'), 25)
    page = request.GET.get('page')
    try:
        profiles_page = paginator.page(page)
    except PageNotAnInteger:
        profiles_page = paginator.page(1)
    except EmptyPage:
        profiles_page = paginator.page(paginator.num_pages)

    context = {
        'profiles': profiles_page,
        'total_profiles': total_profiles,
        'profiles_this_week': profiles_this_week,
        'total_education': total_education,
        'total_experience': total_experience,
        'total_documents': total_documents,
        'complete_profiles': complete_profiles,
        'partial_profiles': partial_profiles,
        'incomplete_profiles': incomplete_profiles,
        'active_page': 'job_profiles',
    }
    return render(request, 'dashboard/job_profiles.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_job_profile_detail(request, pk):
    """Admin: view single user job profile details."""
    profile = get_object_or_404(
        UserProfile.objects.select_related('user').prefetch_related(
            'user__educations', 'user__experiences', 'user__documents',
            'user__application_requests__job',
        ),
        pk=pk
    )
    return render(request, 'dashboard/job_profile_detail.html', {
        'profile': profile,
        'active_page': 'job_profiles',
    })


# ─── Blog Posts Dashboard ────────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_posts(request):
    """Admin dashboard for blog posts."""
    posts = Post.objects.select_related('author', 'category').prefetch_related('tags').order_by('-created_at')

    # Stats
    total_posts = Post.objects.count()
    published_posts = Post.objects.filter(status='published').count()
    draft_posts = Post.objects.filter(status='draft').count()
    total_comments = Comment.objects.count()
    pending_comments = Comment.objects.filter(is_approved=False).count()

    # Filters
    status_filter = request.GET.get('status')
    category_filter = request.GET.get('category')
    if status_filter:
        posts = posts.filter(status=status_filter)
    if category_filter:
        posts = posts.filter(category__slug=category_filter)

    # Pagination
    paginator = Paginator(posts, 20)
    page = request.GET.get('page')
    try:
        posts_page = paginator.page(page)
    except PageNotAnInteger:
        posts_page = paginator.page(1)
    except EmptyPage:
        posts_page = paginator.page(paginator.num_pages)

    categories = Category.objects.all()

    context = {
        'posts': posts_page,
        'categories': categories,
        'total_posts': total_posts,
        'published_posts': published_posts,
        'draft_posts': draft_posts,
        'total_comments': total_comments,
        'pending_comments': pending_comments,
        'active_page': 'posts',
        'status_filter': status_filter,
        'category_filter': category_filter,
    }
    return render(request, 'dashboard/posts.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_post_detail(request, pk):
    """Admin: view single blog post details."""
    post = get_object_or_404(
        Post.objects.select_related('author', 'category').prefetch_related('tags', 'comments'),
        pk=pk
    )
    return render(request, 'dashboard/post_detail.html', {
        'post': post,
        'active_page': 'posts',
    })


# ─── News Subscribers Dashboard ──────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_news_subscribers(request):
    """Admin dashboard for news subscribers."""
    subscribers = NewsSubscriber.objects.prefetch_related('boards').order_by('-subscribed_at')

    # Stats
    total_subscribers = subscribers.count()
    active_subscribers = subscribers.filter(is_active=True).count()
    inactive_subscribers = subscribers.filter(is_active=False).count()

    # Filters
    status_filter = request.GET.get('status')
    if status_filter == 'active':
        subscribers = subscribers.filter(is_active=True)
    elif status_filter == 'inactive':
        subscribers = subscribers.filter(is_active=False)

    # Pagination
    paginator = Paginator(subscribers, 25)
    page = request.GET.get('page')
    try:
        subs_page = paginator.page(page)
    except PageNotAnInteger:
        subs_page = paginator.page(1)
    except EmptyPage:
        subs_page = paginator.page(paginator.num_pages)

    context = {
        'subscribers': subs_page,
        'total_subscribers': total_subscribers,
        'active_subscribers': active_subscribers,
        'inactive_subscribers': inactive_subscribers,
        'active_page': 'news_subscribers',
        'status_filter': status_filter,
    }
    return render(request, 'dashboard/news_subscribers.html', context)


# ─── Blog Post Create/Edit ───────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_post_create(request):
    """Admin: create a new blog post in the custom dashboard."""
    if request.method == 'POST':
        form = PostForm(request.POST, request.FILES)
        if form.is_valid():
            post = form.save(commit=False)
            post.author = request.user
            post.save()
            form.save_m2m()  # Save tags ManyToMany
            messages.success(request, 'Post created successfully.')
            return redirect('dashboard_posts')
    else:
        form = PostForm()
    return render(request, 'dashboard/post_form.html', {
        'form': form,
        'title': 'New Post',
        'active_page': 'posts',
    })


@login_required(login_url='/dashboard/login/')
def dashboard_post_edit(request, pk):
    """Admin: edit an existing blog post in the custom dashboard."""
    post = get_object_or_404(Post, pk=pk)
    if request.method == 'POST':
        form = PostForm(request.POST, request.FILES, instance=post)
        if form.is_valid():
            post = form.save()
            messages.success(request, 'Post updated successfully.')
            return redirect('dashboard_posts')
    else:
        form = PostForm(instance=post)
    return render(request, 'dashboard/post_form.html', {
        'form': form,
        'title': 'Edit Post',
        'active_page': 'posts',
    })


# ─── Service Plans Dashboard ─────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_service_plans(request):
    """Admin dashboard for service plans."""
    plans = ServicePlan.objects.order_by('-created_at')
    total_plans = plans.count()
    active_plans = plans.filter(is_active=True).count()
    total_revenue = sum(p.price for p in plans)

    paginator = Paginator(plans, 20)
    page = request.GET.get('page')
    try:
        plans_page = paginator.page(page)
    except PageNotAnInteger:
        plans_page = paginator.page(1)
    except EmptyPage:
        plans_page = paginator.page(paginator.num_pages)

    return render(request, 'dashboard/service_plans.html', {
        'plans': plans_page,
        'total_plans': total_plans,
        'active_plans': active_plans,
        'total_revenue': total_revenue,
        'active_page': 'service_plans',
    })


# ─── Comments Dashboard ──────────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_comments(request):
    """Admin dashboard for blog post comments."""
    comments_list = Comment.objects.select_related('post').order_by('-created_at')
    total_comments = comments_list.count()
    pending_comments = comments_list.filter(is_approved=False).count()
    approved_comments = comments_list.filter(is_approved=True).count()

    status_filter = request.GET.get('status')
    if status_filter == 'pending':
        comments_list = comments_list.filter(is_approved=False)
    elif status_filter == 'approved':
        comments_list = comments_list.filter(is_approved=True)

    paginator = Paginator(comments_list, 25)
    page = request.GET.get('page')
    try:
        comments_page = paginator.page(page)
    except PageNotAnInteger:
        comments_page = paginator.page(1)
    except EmptyPage:
        comments_page = paginator.page(paginator.num_pages)

    return render(request, 'dashboard/comments.html', {
        'comments': comments_page,
        'total_comments': total_comments,
        'pending_comments': pending_comments,
        'approved_comments': approved_comments,
        'active_page': 'comments',
        'status_filter': status_filter,
    })


# ─── Categories Dashboard ────────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_categories(request):
    """Admin dashboard for blog/news categories."""
    categories = Category.objects.annotate(post_count=Count('posts')).order_by('name')
    total_categories = categories.count()
    blog_categories = categories.filter(type='blog').count()
    news_categories = categories.filter(type='news').count()

    return render(request, 'dashboard/categories.html', {
        'categories': categories,
        'total_categories': total_categories,
        'blog_categories': blog_categories,
        'news_categories': news_categories,
        'active_page': 'categories',
    })


# ─── Tags Dashboard ──────────────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_tags(request):
    """Admin dashboard for blog tags."""
    tags = Tag.objects.annotate(post_count=Count('posts')).order_by('name')
    total_tags = tags.count()

    paginator = Paginator(tags, 30)
    page = request.GET.get('page')
    try:
        tags_page = paginator.page(page)
    except PageNotAnInteger:
        tags_page = paginator.page(1)
    except EmptyPage:
        tags_page = paginator.page(paginator.num_pages)

    return render(request, 'dashboard/tags.html', {
        'tags': tags_page,
        'total_tags': total_tags,
        'active_page': 'tags',
    })


# ─── Activity Logs Dashboard ─────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_activity_logs(request):
    """Admin dashboard for activity logs."""
    logs = ActivityLog.objects.select_related('user').order_by('-created_at')
    total_logs = logs.count()

    paginator = Paginator(logs, 30)
    page = request.GET.get('page')
    try:
        logs_page = paginator.page(page)
    except PageNotAnInteger:
        logs_page = paginator.page(1)
    except EmptyPage:
        logs_page = paginator.page(paginator.num_pages)

    return render(request, 'dashboard/activity_logs.html', {
        'logs': logs_page,
        'total_logs': total_logs,
        'active_page': 'activity_logs',
    })


# ─── AI Usage Dashboard ──────────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_ai_usage(request):
    """Admin dashboard for AI usage tracking."""
    usage = AIUsage.objects.select_related('user').order_by('-date')
    total_records = usage.count()
    today = timezone.now().date()
    today_usage = AIUsage.objects.filter(date=today).aggregate(total=Count('id'))['total'] or 0

    paginator = Paginator(usage, 25)
    page = request.GET.get('page')
    try:
        usage_page = paginator.page(page)
    except PageNotAnInteger:
        usage_page = paginator.page(1)
    except EmptyPage:
        usage_page = paginator.page(paginator.num_pages)

    return render(request, 'dashboard/ai_usage.html', {
        'usage': usage_page,
        'total_records': total_records,
        'today_usage': today_usage,
        'active_page': 'ai_usage',
    })


# ─── Chat Sessions Dashboard ─────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_chat_sessions(request):
    """Admin dashboard for chat sessions."""
    sessions = ChatSession.objects.select_related('user', 'mcq').order_by('-created_at')
    total_sessions = sessions.count()
    active_sessions = sessions.filter(is_active=True).count()

    paginator = Paginator(sessions, 25)
    page = request.GET.get('page')
    try:
        sessions_page = paginator.page(page)
    except PageNotAnInteger:
        sessions_page = paginator.page(1)
    except EmptyPage:
        sessions_page = paginator.page(paginator.num_pages)

    return render(request, 'dashboard/chat_sessions.html', {
        'sessions': sessions_page,
        'total_sessions': total_sessions,
        'active_sessions': active_sessions,
        'active_page': 'chat_sessions',
    })


# ─── Chat Messages Dashboard ─────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_chat_messages(request):
    """Admin dashboard for chat messages."""
    messages_list = ChatMessage.objects.select_related('session').order_by('-created_at')
    total_messages = messages_list.count()
    user_messages = messages_list.filter(role='user').count()
    assistant_messages = messages_list.filter(role='assistant').count()

    paginator = Paginator(messages_list, 30)
    page = request.GET.get('page')
    try:
        messages_page = paginator.page(page)
    except PageNotAnInteger:
        messages_page = paginator.page(1)
    except EmptyPage:
        messages_page = paginator.page(paginator.num_pages)

    return render(request, 'dashboard/chat_messages.html', {
        'chat_messages': messages_page,
        'total_messages': total_messages,
        'user_messages': user_messages,
        'assistant_messages': assistant_messages,
        'active_page': 'chat_messages',
    })


# ─── Django Users Dashboard ──────────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_django_users(request):
    """Admin dashboard for Django auth users."""
    users = User.objects.annotate(
        post_count=Count('posts', distinct=True),
        app_count=Count('application_requests', distinct=True),
    ).order_by('-date_joined')
    total_users = users.count()
    staff_users = users.filter(is_staff=True).count()
    active_users = users.filter(is_active=True).count()

    paginator = Paginator(users, 25)
    page = request.GET.get('page')
    try:
        users_page = paginator.page(page)
    except PageNotAnInteger:
        users_page = paginator.page(1)
    except EmptyPage:
        users_page = paginator.page(paginator.num_pages)

    return render(request, 'dashboard/django_users.html', {
        'users': users_page,
        'total_users': total_users,
        'staff_users': staff_users,
        'active_users': active_users,
        'active_page': 'django_users',
    })


# ─── AI Subscriptions Dashboard ──────────────────────────────────────────────

@login_required(login_url='/dashboard/login/')
def dashboard_ai_subscriptions(request):
    """Admin dashboard for AI plan subscription requests."""
    subs = AISubscription.objects.select_related('user').exclude(plan='free')

    # Stats
    total = subs.count()
    pending = subs.filter(status='pending').count()
    active = subs.filter(status='active').count()
    expired = subs.filter(status='expired').count()
    cancelled = subs.filter(status='cancelled').count()

    # Filters
    status_filter = request.GET.get('status')
    if status_filter:
        subs = subs.filter(status=status_filter)

    # Pagination
    paginator = Paginator(subs.order_by('-created_at'), 25)
    page = request.GET.get('page')
    try:
        subs_page = paginator.page(page)
    except PageNotAnInteger:
        subs_page = paginator.page(1)
    except EmptyPage:
        subs_page = paginator.page(paginator.num_pages)

    context = {
        'subs': subs_page,
        'total': total,
        'pending': pending,
        'active': active,
        'expired': expired,
        'cancelled': cancelled,
        'active_page': 'ai_subscriptions',
        'status_filter': status_filter,
    }
    return render(request, 'dashboard/ai_subscriptions.html', context)


@login_required(login_url='/dashboard/login/')
def dashboard_ai_subscription_approve(request, pk):
    """Admin approves a pending AI subscription — sets status to active."""
    sub = get_object_or_404(AISubscription, pk=pk)
    if sub.status == 'pending':
        sub.status = 'active'
        sub.expires_at = timezone.now() + timedelta(days=30)
        sub.save(update_fields=['status', 'expires_at', 'updated_at'])
        messages.success(request, f'Approved {sub.plan} plan for {sub.user.username}.')
        # Send confirmation email to user
        try:
            from django.core.mail import send_mail
            from django.conf import settings
            plan_name = sub.plan.title()
            send_mail(
                subject=f'ImtihanHub: Your {plan_name} Plan is Active!',
                message=(
                    f'Hello {sub.user.first_name or sub.user.username},\n\n'
                    f'Your {plan_name} plan subscription has been approved and is now active!\n\n'
                    f'Plan: {plan_name}\n'
                    f'Questions per day: {sub.daily_limit}\n'
                    f'Valid until: {sub.expires_at.strftime("%d %B %Y") if sub.expires_at else "N/A"}\n\n'
                    f'You can now enjoy your increased daily AI question limit.\n\n'
                    f'Thank you for choosing ImtihanHub!\n'
                ),
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@imtihanhub.com'),
                recipient_list=[sub.user.email],
                fail_silently=True,
            )
        except Exception:
            pass
    else:
        messages.warning(request, f'Subscription is not pending (current: {sub.status}).')
    return redirect('dashboard_ai_subscriptions')


@login_required(login_url='/dashboard/login/')
def dashboard_ai_subscription_reject(request, pk):
    """Admin rejects a pending AI subscription — sets status to cancelled."""
    sub = get_object_or_404(AISubscription, pk=pk)
    if sub.status == 'pending':
        sub.status = 'cancelled'
        sub.save(update_fields=['status', 'updated_at'])
        messages.success(request, f'Rejected subscription for {sub.user.username}.')
        # Send rejection email to user
        try:
            from django.core.mail import send_mail
            from django.conf import settings
            send_mail(
                subject='ImtihanHub: Subscription Request Update',
                message=(
                    f'Hello {sub.user.first_name or sub.user.username},\n\n'
                    f'Unfortunately, your AI subscription payment could not be verified.\n'
                    f'Your subscription request has been cancelled.\n\n'
                    f'If you believe this is an error, please contact our support team.\n\n'
                    f'Thank you for choosing ImtihanHub!\n'
                ),
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@imtihanhub.com'),
                recipient_list=[sub.user.email],
                fail_silently=True,
            )
        except Exception:
            pass
    else:
        messages.warning(request, f'Subscription is not pending (current: {sub.status}).')
    return redirect('dashboard_ai_subscriptions')
